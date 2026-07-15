"""Extractor: FORMATO ASISTENCIA ERC .xlsm -> DatasetCanonico.

Roster en la hoja Estudiantes + 6 hojas de sesión (Zoom pegado). Matchea email-first,
dedup de minutos por persona y estado por umbrales.
"""

from __future__ import annotations

from openpyxl import load_workbook

from ...domain.dedup import sumar_por_persona
from ...domain.duracion import a_minutos
from ...domain.estados import estado_por_minutos
from ...domain.matcher import matchear
from ...domain.modelos import Asistencia, DatasetCanonico, Nivel, Persona, Sesion

SESIONES = {"1 Viernes": "1V", "1 sábado": "1S", "2 Viernes": "2V",
            "2 sábado": "2S", "3 Viernes": "3V", "3 sábado": "3S"}
_SES_NOMBRE, _SES_DUR, _SES_CORREO = 1, 2, 6


class ErcXlsmExtractor:
    def __init__(self, path: str, *, modulo: str = "", programa: str = "Energías Renovables",
                 cohorte: str = "", presente_min: float = 60, parcial_min: float = 29,
                 umbral_auto: float = 90, umbral_revisar: float = 70) -> None:
        self.path = path
        self.modulo = modulo
        self.programa = programa
        self.cohorte = cohorte
        self.presente_min = presente_min
        self.parcial_min = parcial_min
        self.umbral_auto = umbral_auto
        self.umbral_revisar = umbral_revisar

    def _roster(self, wb) -> list[Persona]:
        ws = wb["Estudiantes"]
        personas: list[Persona] = []
        for r in range(2, ws.max_row + 1):
            ap = ws.cell(row=r, column=1).value
            nom = ws.cell(row=r, column=2).value
            comp = ws.cell(row=r, column=4).value
            cor = ws.cell(row=r, column=3).value
            if (ap is None or str(ap).strip() == "") and (comp is None or str(comp).strip() == ""):
                break
            nombre = str(comp).strip() if comp else f"{ap or ''}, {nom or ''}"
            personas.append(Persona(nombre=nombre, correo=str(cor or "").strip(),
                                    programa=self.programa, cohorte=self.cohorte))
        return personas

    def extraer(self) -> DatasetCanonico:
        wb = load_workbook(self.path, data_only=True, keep_vba=True)
        personas = self._roster(wb)
        mod = self.modulo or "ERC"
        sesiones: list[Sesion] = []
        asistencias: list[Asistencia] = []
        sin_match: list[dict] = []

        for hoja, etq in SESIONES.items():
            if hoja not in wb.sheetnames:
                continue
            ws = wb[hoja]
            sesion = Sesion(id=f"erc:{mod}:{etq}", programa=self.programa, cohorte=self.cohorte,
                            modulo=mod, etiqueta=etq)
            sesiones.append(sesion)
            pares: list[tuple[Persona, float]] = []
            for r in range(2, ws.max_row + 1):
                nombre = ws.cell(row=r, column=_SES_NOMBRE).value
                if nombre is None or str(nombre).strip() == "":
                    continue
                minutos = a_minutos(ws.cell(row=r, column=_SES_DUR).value)
                correo = str(ws.cell(row=r, column=_SES_CORREO).value or "").strip()
                m = matchear(str(nombre).strip(), correo, personas, self.umbral_auto, self.umbral_revisar)
                if m.nivel is Nivel.AUTO and m.persona is not None:
                    pares.append((m.persona, minutos))
                else:
                    sin_match.append({"sesion": etq, "nombre": str(nombre).strip(),
                                      "correo": correo, "minutos": round(minutos, 1),
                                      "nivel": m.nivel.value})
            mins = sumar_por_persona(pares)
            for p in personas:
                mm = mins.get(p.clave, 0.0)
                asistencias.append(Asistencia(
                    persona_id=p.id, sesion_id=sesion.id,
                    estado=estado_por_minutos(mm, self.presente_min, self.parcial_min),
                    minutos=round(mm, 1), origen="erc-xlsm", nombre_origen=p.nombre))

        return DatasetCanonico(fuente="erc-xlsm", personas=personas, sesiones=sesiones,
                               asistencias=asistencias, sin_match=sin_match)

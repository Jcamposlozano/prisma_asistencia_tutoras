"""Extractor: carpeta de cohorte (Listado.xlsx + N CSV por clase) -> DatasetCanonico.

La fecha de cada clase sale del NOMBRE del archivo (Clase <día> <mes> <año>). Las
personas vienen del Listado (columna A). Cada CSV es Zoom -> match por nombre.
"""

from __future__ import annotations

import datetime as _dt
import glob
import os
import re

from openpyxl import load_workbook

from ...domain.dedup import sumar_por_persona
from ...domain.estados import estado_por_minutos
from ...domain.matcher import matchear
from ...domain.modelos import Asistencia, DatasetCanonico, Nivel, Persona, Sesion
from ._zoom_common import clasificar, leer_csv_zoom

_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
          "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
          "noviembre": 11, "diciembre": 12}
_ABBR = {1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun", 7: "jul",
         8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"}
_RE_FECHA = re.compile(r"(\d{1,2})\s+([a-záéíóúñ]+)\s+(\d{4})", re.I)


def _fecha_nombre(path: str) -> _dt.date | None:
    m = _RE_FECHA.search(os.path.basename(path).lower())
    if not m:
        return None
    mes = _MESES.get(m.group(2))
    if not mes:
        return None
    try:
        return _dt.date(int(m.group(3)), mes, int(m.group(1)))
    except ValueError:
        return None


def _etiqueta(f: _dt.date) -> str:
    return f"{f.day:02d} {_ABBR[f.month]}"


class MatrizDirExtractor:
    def __init__(self, carpeta: str, *, listado: str, hoja: str | None = None,
                 programa: str = "", presente_min: float = 60, parcial_min: float = 29,
                 umbral_auto: float = 90, umbral_revisar: float = 70) -> None:
        self.carpeta = carpeta
        self.listado = listado
        self.hoja = hoja or os.path.basename(os.path.normpath(carpeta))
        self.programa = programa
        self.presente_min = presente_min
        self.parcial_min = parcial_min
        self.umbral_auto = umbral_auto
        self.umbral_revisar = umbral_revisar

    def _personas(self) -> list[Persona]:
        wb = load_workbook(self.listado, read_only=True, data_only=True)
        try:
            if self.hoja not in wb.sheetnames:
                raise KeyError(f"La hoja {self.hoja!r} no existe en {self.listado}")
            ws = wb[self.hoja]
            out: list[Persona] = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                v = row[0]
                if v is None or str(v).strip() == "":
                    break
                out.append(Persona(nombre=str(v).replace("\xa0", " ").strip(),
                                   cohorte=self.hoja, programa=self.programa))
            return out
        finally:
            wb.close()

    def extraer(self) -> DatasetCanonico:
        personas = self._personas()
        csvs = sorted(glob.glob(os.path.join(self.carpeta, "*.csv")),
                      key=lambda p: _fecha_nombre(p) or _dt.date.min)
        sesiones: list[Sesion] = []
        asistencias: list[Asistencia] = []
        sin_match: list[dict] = []
        filtrados: list[dict] = []

        for csvp in csvs:
            fecha = _fecha_nombre(csvp)
            fiso = fecha.isoformat() if fecha else ""
            etq = _etiqueta(fecha) if fecha else os.path.splitext(os.path.basename(csvp))[0]
            sesion = Sesion(id=f"matriz:{self.hoja}:{fiso or etq}", cohorte=self.hoja,
                            programa=self.programa, etiqueta=etq, fecha=fiso)
            sesiones.append(sesion)
            pares: list[tuple[Persona, float]] = []
            for f in leer_csv_zoom(csvp):
                cat, nm = clasificar(f["nombre"], f["invitado"], f["correo"])
                if cat in ("HOST", "BOT") or (cat == "DISPOSITIVO" and not nm):
                    filtrados.append({"sesion": etq, "nombre": f["nombre"], "categoria": cat})
                    continue
                m = matchear(nm, f["correo"], personas, self.umbral_auto, self.umbral_revisar)
                if m.nivel is Nivel.AUTO and m.persona is not None:
                    pares.append((m.persona, f["minutos"]))
                else:
                    sin_match.append({"sesion": etq, "nombre": f["nombre"],
                                      "minutos": round(f["minutos"], 1), "nivel": m.nivel.value})
            mins = sumar_por_persona(pares)
            for p in personas:
                mm = mins.get(p.clave, 0.0)
                asistencias.append(Asistencia(
                    persona_id=p.id, sesion_id=sesion.id,
                    estado=estado_por_minutos(mm, self.presente_min, self.parcial_min),
                    minutos=round(mm, 1), origen="matriz", nombre_origen=p.nombre))

        return DatasetCanonico(fuente="matriz", personas=personas, sesiones=sesiones,
                               asistencias=asistencias, sin_match=sin_match, filtrados=filtrados)

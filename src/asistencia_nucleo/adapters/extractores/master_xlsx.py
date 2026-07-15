"""Extractor: Asistencia Máster .xlsx (grid por cohorte) -> DatasetCanonico.

Columna A = personas; columnas de sesión = encabezados tipo FECHA (los años están
corruptos, pero el TIPO fecha las identifica); marcas X / vacío / EXCUSA. Ya viene
vinculado (no requiere matching con Zoom).
"""

from __future__ import annotations

import datetime as _dt

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ...domain.modelos import Asistencia, DatasetCanonico, EstadoCanonico, Persona, Sesion
from ...domain.normalizar import normalizar, quitar_diacriticos


def _estado_marca(v) -> EstadoCanonico:
    if v is None or str(v).strip() == "":
        return EstadoCanonico.AUSENTE
    t = quitar_diacriticos(str(v)).strip().lower()
    if t == "x":
        return EstadoCanonico.PRESENTE
    if "excus" in t or "incapacid" in t:
        return EstadoCanonico.EXCUSA
    return EstadoCanonico.DESCONOCIDO


class MasterXlsxExtractor:
    def __init__(self, path: str, *, hoja: str | None = None, programa: str = "Máster") -> None:
        self.path = path
        self.hoja = hoja
        self.programa = programa

    def extraer(self) -> DatasetCanonico:
        wb = load_workbook(self.path, data_only=True)
        hojas = [self.hoja] if self.hoja else list(wb.sheetnames)
        personas: list[Persona] = []
        sesiones: list[Sesion] = []
        asistencias: list[Asistencia] = []

        for h in hojas:
            if h not in wb.sheetnames:
                raise KeyError(f"La hoja {h!r} no existe. Hojas: {wb.sheetnames}")
            ws = wb[h]
            cols_sesion = [
                c for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(row=1, column=c).value, (_dt.datetime, _dt.date))
            ]
            sesion_de: dict[int, Sesion] = {}
            for c in cols_sesion:
                s = Sesion(id=f"master:{h}:{get_column_letter(c)}", cohorte=h,
                           programa=self.programa, etiqueta=get_column_letter(c))
                sesiones.append(s)
                sesion_de[c] = s

            for r in range(2, ws.max_row + 1):
                nombre = ws.cell(row=r, column=1).value
                if nombre is None or str(nombre).strip() == "":
                    break
                nombre = str(nombre).replace("\xa0", " ").strip()
                p = Persona(nombre=nombre, cohorte=h, programa=self.programa,
                            id=f"master:{h}|{normalizar(nombre).replace(' ', '-')}")
                personas.append(p)
                for c in cols_sesion:
                    asistencias.append(Asistencia(
                        persona_id=p.id, sesion_id=sesion_de[c].id,
                        estado=_estado_marca(ws.cell(row=r, column=c).value),
                        origen="master", nombre_origen=nombre))

        return DatasetCanonico(fuente="master", personas=personas, sesiones=sesiones,
                               asistencias=asistencias)

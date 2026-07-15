"""Renderer 'consolidado': matriz Persona × sesión + conteos y % de asistencia."""

from __future__ import annotations

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ...domain.estandar import indice_asistencias
from ...domain.modelos import EstadoCanonico, Estandar

log = logging.getLogger(__name__)

_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")
_HEADER = PatternFill("solid", fgColor="D9D9D9")
_CODIGO = {
    EstadoCanonico.PRESENTE: ("P", PatternFill("solid", fgColor="C6EFCE")),
    EstadoCanonico.PARCIAL: ("D", PatternFill("solid", fgColor="FFEB9C")),
    EstadoCanonico.AUSENTE: ("o", PatternFill("solid", fgColor="FFC7CE")),
    EstadoCanonico.EXCUSA: ("E", PatternFill("solid", fgColor="BDD7EE")),
    EstadoCanonico.DESCONOCIDO: ("?", PatternFill("solid", fgColor="D0D0D0")),
}


class ConsolidadoXlsxRenderer:
    def render(self, estandar: Estandar, salida: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "consolidado"

        personas = sorted(estandar.personas, key=lambda p: p.nombre.lower())
        # Orden por fecha si existe; si no, se conserva el orden natural de extracción.
        orden = {s.id: i for i, s in enumerate(estandar.sesiones)}
        sesiones = sorted(estandar.sesiones, key=lambda s: (s.fecha or "", orden[s.id]))
        idx = indice_asistencias(estandar)
        n = len(sesiones)

        ws["A1"] = "Leyenda: P=presente  D=parcial  o=ausente  E=excusa  ?=desconocido"
        ws["A1"].font = _BOLD
        fila_cab = 2
        cab = ["Persona", *[s.etiqueta for s in sesiones],
               "# Pres", "# Parc", "# Aus", "# Exc", "% Asist"]
        for j, h in enumerate(cab, start=1):
            c = ws.cell(row=fila_cab, column=j, value=h)
            c.font = _BOLD
            c.fill = _HEADER
            c.alignment = _CENTER

        for i, p in enumerate(personas, start=fila_cab + 1):
            ws.cell(row=i, column=1, value=p.nombre)
            cont = {e: 0 for e in EstadoCanonico}
            for j, s in enumerate(sesiones, start=2):
                a = idx.get((p.id, s.id))
                estado = a.estado if a else EstadoCanonico.AUSENTE
                cont[estado] += 1
                codigo, fill = _CODIGO[estado]
                cell = ws.cell(row=i, column=j, value=codigo)
                cell.alignment = _CENTER
                cell.fill = fill
            ws.cell(row=i, column=2 + n, value=cont[EstadoCanonico.PRESENTE]).alignment = _CENTER
            ws.cell(row=i, column=3 + n, value=cont[EstadoCanonico.PARCIAL]).alignment = _CENTER
            ws.cell(row=i, column=4 + n, value=cont[EstadoCanonico.AUSENTE]).alignment = _CENTER
            ws.cell(row=i, column=5 + n, value=cont[EstadoCanonico.EXCUSA]).alignment = _CENTER
            pct = round(cont[EstadoCanonico.PRESENTE] / n * 100, 1) if n else 0.0
            ws.cell(row=i, column=6 + n, value=pct).alignment = _CENTER

        ws.column_dimensions["A"].width = 36
        for j in range(2, 2 + n):
            ws.column_dimensions[get_column_letter(j)].width = 6
        ws.freeze_panes = ws.cell(row=fila_cab + 1, column=2).coordinate

        wb.save(salida)
        log.info("Consolidado escrito en %s (%d personas x %d sesiones)", salida, len(personas), n)

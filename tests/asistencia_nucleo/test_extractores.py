"""Integración de los extractores sobre muestras reales (skip si faltan)."""

from __future__ import annotations

import openpyxl

from asistencia_nucleo.adapters.extractores.erc_xlsm import ErcXlsmExtractor
from asistencia_nucleo.adapters.extractores.master_xlsx import MasterXlsxExtractor
from asistencia_nucleo.adapters.extractores.matriz_dir import MatrizDirExtractor
from asistencia_nucleo.adapters.extractores.zoom_csv import ZoomCsvExtractor
from asistencia_nucleo.adapters.salida.consolidado_xlsx import ConsolidadoXlsxRenderer
from asistencia_nucleo.domain.estandar import merge
from asistencia_nucleo.domain.modelos import EstadoCanonico, Estandar
from .conftest import (
    ERC,
    LISTADO,
    MASTER,
    MATRIZ_DIR,
    ZOOM,
    requiere_erc,
    requiere_master,
    requiere_matriz,
    requiere_zoom,
)


@requiere_erc
def test_erc_cross_check_con_asistencia_erc():
    ds = ErcXlsmExtractor(ERC, modulo="ERC10").extraer()
    assert len(ds.personas) == 38
    assert len(ds.sesiones) == 6
    assert len(ds.asistencias) == 228
    pres = {
        s.etiqueta: sum(1 for a in ds.asistencias
                        if a.sesion_id == s.id and a.estado is EstadoCanonico.PRESENTE)
        for s in ds.sesiones
    }
    # mismos conteos A/sesión validados en el proyecto asistencia-erc (99.1% vs Total humano)
    assert pres == {"1V": 20, "1S": 16, "2V": 16, "2S": 16, "3V": 12, "3S": 13}


@requiere_zoom
def test_zoom_extractor():
    ds = ZoomCsvExtractor(ZOOM).extraer()
    assert len(ds.sesiones) == 1
    assert ds.sesiones[0].fecha == "2026-04-28"
    assert len(ds.filtrados) >= 2  # host + bot (Fathom)
    assert any(a.estado is EstadoCanonico.PRESENTE for a in ds.asistencias)


@requiere_matriz
def test_matriz_extractor():
    ds = MatrizDirExtractor(MATRIZ_DIR, listado=LISTADO).extraer()
    assert len(ds.personas) == 52
    assert len(ds.sesiones) == 10
    assert len(ds.asistencias) == 520


@requiere_master
def test_master_extractor():
    ds = MasterXlsxExtractor(MASTER, hoja="2025-2 D.Business").extraer()
    assert len(ds.personas) == 52
    assert len(ds.sesiones) == 49
    assert sum(1 for a in ds.asistencias if a.estado is EstadoCanonico.PRESENTE) > 0


@requiere_erc
def test_e2e_extraer_json_render(tmp_path):
    # extraer -> estándar -> JSON round-trip -> consolidado
    ds = ErcXlsmExtractor(ERC, modulo="ERC10").extraer()
    est = merge([ds], generado="2026-07-09")
    est2 = Estandar.from_json(est.to_json())
    assert est2.meta["n_asistencias"] == 228

    salida = str(tmp_path / "consolidado.xlsx")
    ConsolidadoXlsxRenderer().render(est2, salida)
    wb = openpyxl.load_workbook(salida)
    assert "consolidado" in wb.sheetnames
    ws = wb["consolidado"]
    # encabezado de sesión presente (fila 2, col 2)
    assert ws.cell(row=2, column=2).value == "1V"
